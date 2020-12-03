
# bot.py
import os
import random
from discord.ext import commands
from dotenv import load_dotenv
from bs4 import BeautifulSoup as bs
import requests
import re
import shutil
import os
import time
import MySQLdb
import logging
from urllib.parse import urlparse
from Educar import database, requests_core, saber_fix
from time import time
from functools import partial
from concurrent.futures import ProcessPoolExecutor
import concurrent.futures
from concurrent import futures
import shutil
import pymysql
import openpyxl
import json
import discord
import calendar
import time

load_dotenv()
TOKEN = 'NzUxNTE3MjcyMjcwNzAwNjM1.X1KPAA.zrRT6FhykVmrXvPUxhh3xftOBqI'

bot = commands.Bot(command_prefix='!')


@bot.command(name='99', help='Frase para probar el bot //ingles//')
async def nine_nine(ctx):
    brooklyn_99_quotes = [
        'I\'m the human form of the 💯 emoji.',
        'Bingpot!',
        (
            'Cool. Cool cool cool cool cool cool cool, '
            'no doubt no doubt no doubt no doubt.'
        ),
    ]

    response = random.choice(brooklyn_99_quotes)
    await ctx.send(response)


@bot.command(name='numeros', help='Simulates rolling dice.')
async def roll(ctx, number_of_dice: int, number_of_sides: int):
    dice = [
        str(random.choice(range(1, number_of_sides + 1)))
        for _ in range(number_of_dice)
    ]
    doce = 'Usuario reparado con normalidad'
    await ctx.send(', '.join(dice))
    await ctx.send(doce)


@bot.command(name='usuario', help='Crea y matricula un usuario en el nuevo AULA')
async def usuario(ctx, usuario: str, password: str):
    mysql = MySQLdb.connect(host='educar2020db.c6r9vkodxz44.us-east-1.rds.amazonaws.com',
                        user="EducarDB",
                        passwd="Db2020!",
                        db="global_sie")
    mysql2 = MySQLdb.connect(host='educar2020db.c6r9vkodxz44.us-east-1.rds.amazonaws.com',
                         user="EducarDB",
                         passwd="Db2020!",
                         db="2018_chm_lms_demo")
    mysql3 = MySQLdb.connect(host='db-produccion-educarsie2-serveless.cluster-cxlnmcx1f2mz.us-east-2.rds.amazonaws.com',
                         user="superadmin_prod",
                         passwd="%Educ4r2020%",
                         db="2018_chm_lms_demo")

    URL = 'https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2F'
    LOGIN_ROUTE = ''

    HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
               'origin': URL, 'referer': URL + LOGIN_ROUTE}

    s = requests.session()
    soup = bs(s.get(
        'https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2Fsie%2F').text, 'html.parser')
    soup.find('input', {"name": "lt"})
    sp = soup.find('input', {"name": "lt"})
    cas_token = sp['value']

    user1 = usuario
    password1 = password

    login_payload = {
        'lt': cas_token,
        'execution': 'e1s1',
        '_eventId': 'submit',
        'username': user1,
        'password': password1,
        'submit': 'ENTRAR'
    }

    login_req = s.post(URL, headers=HEADERS, data=login_payload)

    cookies = login_req.cookies

    r = s.get('http://sie.educar.com.co/Saber2019/')
    rm = s.post('http://sie.educar.com.co/Saber2019/')
    soup = bs(s.get(
        'http://sie.educar.com.co/PruebasSaber/mod/quiz/view.php?id=20').text, 'html.parser')

    soup.find('p', {"class": "text-right"})
    sp = soup.find('p', {"class": "text-right"})
    sopa = soup.find_all('p', {"class": "text-right"})
    cursos = sopa[2]
    instituciones = sopa[1]
    for htmls in cursos:

        curso = htmls

    for htmls in instituciones:

        institucion = htmls

    cur1 = mysql.cursor()
    cur2 = mysql2.cursor()
    cur3 = mysql3.cursor()
    cur4 = mysql2.cursor()
    cur1.execute(
        'SELECT id FROM instituciones WHERE nombre=%s', (institucion,))
    cat = cur1.fetchone()
    print('INSTITUCION=', institucion, 'codigo=', cat)
    response1 = 'INSTITUCION=', institucion, 'codigo=', cat
    await ctx.send(response1)
    curso = '%' + curso + '%'
    cur2.execute('SELECT code FROM course WHERE category_code = %s AND title LIKE %s',
                 (cat, curso))
    curso_id = cur2.fetchall()

    cur4.execute(
        'SELECT user_id FROM user WHERE username=%s', (usuario,))
    user_idx = cur4.fetchall()

    print('USER ID = ', user_idx)
    response2 = 'USER ID = ', user_idx
    await ctx.send(response2)
    # ya estan los datos, ahora nos logueamos como admin para inscribirlo

    URL = 'https://aula.educar.com.co/'
    LOGIN_ROUTE = ''

    HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
               'origin': URL, 'referer': URL + LOGIN_ROUTE}

    s = requests.session()

    """ csrf_token = s.get(URL).cookies['csrftoken'] """
    user = 'admin'
    password = '$%EDUC4R2020.it'

    login_payload = {
        'login': user,
        'password': password,

    }

    login_req = s.post(URL, headers=HEADERS, data=login_payload)

    print(login_req.status_code)
    await ctx.send(login_req.status_code)

    cookies = login_req.cookies

    soup3 = bs(
        s.get('https://aula.educar.com.co/main/admin/user_add.php').text, 'html.parser')

    soup3.find('input', {"name": "sec_token"})
    sp3 = soup3.find('input', {"name": "sec_token"})
    sec_token = sp3['value']
    print(sec_token)
    await ctx.send(sec_token)
    cookies = login_req.cookies
    curfetch = mysql2.cursor()
    curfetch2 = mysql2.cursor()
    curfetch3 = mysql2.cursor()
    curfetch4 = mysql2.cursor()
    curfetch.execute(
        'SELECT email FROM user WHERE username=%s', (user1,))
    email = curfetch.fetchall()
    curfetch2.execute(
        'SELECT lastname FROM user WHERE username=%s', (user1,))
    apellido = curfetch2.fetchall()
    curfetch3.execute(
        'SELECT firstname FROM user WHERE username=%s', (user1,))
    nombre = curfetch3.fetchall()
    apellido = ''.join(map(str, apellido))
    apellido = apellido.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    nombre = ''.join(map(str, nombre))
    nombre = nombre.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    email = ''.join(map(str, email))
    email = email.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    usuario_payload = {
        'lastname': nombre,
        'firstname': apellido,
        'official_code': '',
        'email': email,
        'phone': '',

        'username': user1,
        'password[password_auto]': '0',
        'password[password]': '$%EDUC4R2020.it',
        'status': '5',
        'admin[platform_admin]': '0',
        'language': 'spanish2',
        'mail[send_mail]': '0',
        'radio_expiration_date': '0',
        'expiration_date': '2030-08-31 20:29',
        'active': '1',
        'extra_legal_accept': '',
        'extra_already_logged_in': '',
        'extra_update_type': '',
        'extra_rssfeeds': '',
        'extra_dashboard': '',
        'extra_timezone': '',
        'extra_mail_notify_invitation': '1',
        'extra_mail_notify_group_message': '1',
        'extra_user_chat_status': '',
        'extra_google_calendar_url': '',
        'extra_captcha_blocked_until_date': '',
        'extra_skype': '',
        'extra_linkedin_url': '',
        'extra_mail_notify_message': '1',
        'extra_request_for_legal_agreement_consent_removal_justification': '',
        'extra_request_for_delete_account_justification': '',
        'extra_request_for_legal_agreement_consent_removal': '',
        'extra_request_for_delete_account': '',
        'submit': '',
        '_qf__user_add': '',
        'item_id': '0',
        'sec_token': sec_token,
    }
    user_req = s.post('https://aula.educar.com.co/main/admin/user_add.php',
                      headers=HEADERS, data=usuario_payload)
    varinfo = nombre, apellido
    await ctx.send(', '.join(varinfo))
    cur3.execute(
        'SELECT user_id FROM user WHERE username=%s', (user1,))
    user_id = cur3.fetchall()
    print(cur3._last_executed)
    print('USER ID = ', user_id)
    response3 = 'USER ID = ', user_id
    await ctx.send(response3)
    ts = time()
    for curso in curso_id:
        curso = ''.join(map(str, curso))
        curso = curso.replace('(', '').replace(
            "'", "").replace(',', '').replace(')', '')
        user_id = ''.join(map(str, user_id))
        user_id = user_id.replace(
            '(', '').replace(',', '').replace(')', '')
        inscribir = 'https://aula.educar.com.co/main/user/subscribe_user.php?cidReq={}&id_session=0&gidReq=0&gradebook=0&origin=&register=yes&type=5&user_id={}'.format(
            curso, user_id)
        s.get(inscribir)

        print('usuario : ', user1,
              'Inscrito satisfactoriamente en el curso', curso)
        response4 = 'usuario : ', user1, 'Inscrito satisfactoriamente en el curso', curso
        await ctx.send(response4)
    tiempo = 'Tardo %s', time() - ts
    await ctx.send(tiempo)
    revisar = 'https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2FAprende%2F&logueo=true&user={}&pass={}'.format(
        user1, password1)
    revisar = ''.join(map(str, revisar))
    revise = 'Revisalo con un click:'
    await ctx.send(revise)
    await ctx.send(revisar)
    mysql.close()
    mysql2.close()
    mysql3.close()

@bot.command(name='profesores', help='Crea profesores')
async def roll(ctx, nombre: str, apellido: str, usuario: str):
     db = database.Database()
     rq = requests_core.Requests()
     crear = rq.crear_profesor(nombre, apellido, usuario)
     await ctx.send(', '.join(crear))

@bot.command(name='matricula', help='creacion de materias & matricula de profesores')
async def roll(ctx, institucion: str, firstname: str, curso: str, grupo: str, usuario: str):
     firstname = firstname.replace('-',' ')
     db = database.Database()
     rq = requests_core.Requests()
     crear = rq.cursos_y_matricula(institucion, firstname, curso, grupo, usuario)
     await ctx.send(', '.join(crear))

@bot.command(name='inscribirprof', help='inscribir profesores en los cursos en los que ya estaban (aula viejo)')
async def prof(ctx, usuario: str, password: str):
     ts = time()
     db = database.Database()
     rq = requests_core.Requests()
     inscribir = rq.matricular_vn(usuario, password)
     await ctx.send(', '.join(inscribir))

@bot.command(name='sie', help='ajustar botones aula SIE')
async def prof(ctx, institucion: str):
     iniciando = 'Ajustando parametros SIE para la institucion......'
     await ctx.send(iniciando)
     db = database.Database()
     rq = requests_core.Requests()
     ajustar = db.quitar_botones(institucion)
     await ctx.send(', '.join(ajustar))


@bot.command(name='tusuario', help='Crea y matricula un usuario en el nuevo AULA MULTITHREADING OPTIMIZADO')
async def usuario(ctx, usuario: str, password: str):

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    mysql = MySQLdb.connect(host='educar2020db.c6r9vkodxz44.us-east-1.rds.amazonaws.com',
                            user="EducarDB",
                            passwd="Db2020!",
                            db="global_sie")
    mysql2 = MySQLdb.connect(host='educar2020db.c6r9vkodxz44.us-east-1.rds.amazonaws.com',
                            user="EducarDB",
                            passwd="Db2020!",
                            db="2018_chm_lms_demo")
    mysql3 = MySQLdb.connect(host='db-produccion-educarsie2-serveless.cluster-cxlnmcx1f2mz.us-east-2.rds.amazonaws.com',
                            user="superadmin_prod",
                            passwd="%Educ4r2020%",
                            db="2018_chm_lms_demo")
    logger = logging.getLogger('try_Log')

    URL = 'https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2F'
    LOGIN_ROUTE = ''

    HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
            'origin': URL, 'referer': URL + LOGIN_ROUTE}

    s = requests.session()
    soup = bs(s.get('https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2Fsie%2F').text, 'html.parser')
    soup.find('input', {"name": "lt"})
    sp = soup.find('input', {"name": "lt"})
    cas_token = sp['value']


    user1 = usuario
    password1 = password


    login_payload = {
        'lt': cas_token,
        'execution': 'e1s1',
        '_eventId': 'submit',
        'username': user1,
        'password': password1,
        'submit': 'ENTRAR'
    }

    login_req = s.post(URL, headers=HEADERS, data=login_payload)


    cookies = login_req.cookies

    r = s.get('http://sie.educar.com.co/Saber2019/')
    rm = s.post('http://sie.educar.com.co/Saber2019/')
    soup = bs(s.get(
        'http://sie.educar.com.co/PruebasSaber/mod/quiz/view.php?id=20').text, 'html.parser')

    soup.find('p', {"class": "text-right"})
    sp = soup.find('p', {"class": "text-right"})
    sopa = soup.find_all('p', {"class": "text-right"})
    cursos = sopa[2]
    instituciones = sopa[1]
    for htmls in cursos:

        curso = htmls

    for htmls in instituciones:

        institucion = htmls

    cur1 = mysql.cursor()
    cur2 = mysql2.cursor()
    cur3 = mysql3.cursor()
    cur4 = mysql2.cursor()
    cur1.execute('SELECT id FROM instituciones WHERE nombre=%s', (institucion,))
    cat = cur1.fetchone()
    print('INSTITUCION=', institucion, 'codigo=', cat)
    data1 = 'INSTITUCION=', institucion, 'codigo=', cat
    await ctx.send(data1)
    curso = '%' + curso + '%'
    cur2.execute('SELECT code FROM course WHERE category_code = %s AND title LIKE %s',
                (cat, curso))
    curso_id = cur2.fetchall()


    cur4.execute('SELECT user_id FROM user WHERE username=%s', (usuario,))
    user_idx = cur4.fetchall()
    print('USER ID = ', user_idx)
    data2 = 'USER ID = ', user_idx
    await ctx.send(data2)
    # ya estan los datos, ahora nos logueamos como admin para inscribirlo

    URL = 'https://aula.educar.com.co/'
    LOGIN_ROUTE = ''


    HEADERS = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.88 Safari/537.36',
            'origin': URL, 'referer': URL + LOGIN_ROUTE}

    s = requests.session()

    """ csrf_token = s.get(URL).cookies['csrftoken'] """
    user = 'admin'
    password = '$%EDUC4R2020.it'

    login_payload = {
        'login': user,
        'password': password,

    }

    login_req = s.post(URL, headers=HEADERS, data=login_payload)

    print(login_req.status_code)

    cookies = login_req.cookies

    soup3 = bs(
        s.get('https://aula.educar.com.co/main/admin/user_add.php').text, 'html.parser')

    soup3.find('input', {"name": "sec_token"})
    sp3 = soup3.find('input', {"name": "sec_token"})
    sec_token = sp3['value']
    print(sec_token)
    cookies = login_req.cookies
    curfetch = mysql2.cursor()
    curfetch2 = mysql2.cursor()
    curfetch3 = mysql2.cursor()
    curfetch4 = mysql2.cursor()
    curfetch.execute('SELECT email FROM user WHERE username=%s', (user1,))
    email = curfetch.fetchall()
    curfetch2.execute('SELECT lastname FROM user WHERE username=%s', (user1,))
    apellido = curfetch2.fetchall()
    curfetch3.execute('SELECT firstname FROM user WHERE username=%s', (user1,))
    nombre = curfetch3.fetchall()
    apellido = ''.join(map(str, apellido))
    apellido = apellido.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    nombre = ''.join(map(str, nombre))
    nombre = nombre.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    email = ''.join(map(str, email))
    email = email.replace('(', '').replace(
        "'", "").replace(',', '').replace(')', '')
    print('datos de usuario', nombre, apellido)
    data3 = 'datos de usuario', nombre, apellido
    await ctx.send(data3)
    usuario_payload = {
        'lastname': apellido,
        'firstname': nombre,
        'official_code': '',
        'email': email,
        'phone': '',
        'username': user1,
        'password[password_auto]': '0',
        'password[password]': '$%EDUC4R2020.it',
        'status': '5',
        'admin[platform_admin]': '0',
        'language': 'spanish2',
        'mail[send_mail]': '0',
        'radio_expiration_date': '0',
        'expiration_date': '2030-08-31 20:29',
        'active': '1',
        'extra_legal_accept': '',
        'extra_already_logged_in': '',
        'extra_update_type': '',
        'extra_rssfeeds': '',
        'extra_dashboard': '',
        'extra_timezone': '',
        'extra_mail_notify_invitation': '1',
        'extra_mail_notify_group_message': '1',
        'extra_user_chat_status': '',
        'extra_google_calendar_url': '',
        'extra_captcha_blocked_until_date': '',
        'extra_skype': '',
        'extra_linkedin_url': '',
        'extra_mail_notify_message': '1',
        'extra_request_for_legal_agreement_consent_removal_justification': '',
        'extra_request_for_delete_account_justification': '',
        'extra_request_for_legal_agreement_consent_removal': '',
        'extra_request_for_delete_account': '',
        'submit': '',
        '_qf__user_add': '',
        'item_id': '0',
        'sec_token': sec_token,
    }
    user_req = s.post('https://aula.educar.com.co/main/admin/user_add.php',
                    headers=HEADERS, data=usuario_payload)

    cur3.execute('SELECT user_id FROM user WHERE username=%s', (user1,))
    user_id = cur3.fetchall()
    print('USER ID = ', user_id)
    data4 = 'USER ID = ', user_id
    await ctx.send(data4)


    URLS = []
    for curso in curso_id:
        curso = ''.join(map(str, curso))
        curso = curso.replace('(', '').replace(
                "'", "").replace(',', '').replace(')', '')
        user_id = ''.join(map(str, user_id))
        user_id = user_id.replace('(', '').replace(',', '').replace(')', '')
        inscribir = 'https://aula.educar.com.co/main/user/subscribe_user.php?cidReq={}&id_session=0&gidReq=0&gradebook=0&origin=&register=yes&type=5&user_id={}'.format(
                curso, user_id)

        apendar = URLS.append(inscribir)


    print(URLS)
    data5 = 'Inscribiendo usuario en los cursos......'
    await ctx.send(data5)
    ts = time()

    with futures.ThreadPoolExecutor(max_workers=10) as executor: ## you can increase the amount of workers, it would increase the amount of thread created
        res = executor.map(s.get,URLS)
    responses = list(res) ## the future is returning a generator. You may want to turn it to list.

    print(responses)
    logging.info('Tomo %s', time() - ts)
    tiempo = 'Tardo %s', time() - ts
    respuesta = 'Usuario inscrito correctamente', tiempo
    await ctx.send(respuesta)
    revisar = 'https://sie.educar.com.co/cas/login?service=http%3A%2F%2Fsie.educar.com.co%2FAprende%2F&logueo=true&user={}&pass={}'.format(
        user1, password1)
    revisar = ''.join(map(str, revisar))
    revise = 'Revisalo con un click:'
    await ctx.send(revise)
    await ctx.send(revisar)
    mysql.close()
    mysql2.close()
    mysql3.close()

@bot.command(name='sbkill', help='Arreglar matricula SABER')
async def prof(ctx, institucion: str):
     iniciando = 'Ajustando parametros SABER para la institucion......'
     await ctx.send(iniciando)
     sb = saber_fix.Saber()

     ajustar = sb.saber_kill(institucion)
     await ctx.send(', '.join(ajustar))

@bot.command(name='genpines', help='Generacion de pines en excel')
async def prof(ctx, institucion: str, cargas: str, numarchivos: int, nom_institucion: str, ciudad: str):
    r = requests

    mysql = pymysql.connect(host='db-produccion-educarsie2-serveless.cluster-cxlnmcx1f2mz.us-east-2.rds.amazonaws.com',
                            user="superadmin_prod",
                            passwd="%Educ4r2020%",
                            db="global_sie")

    HEADERS = {'Content-Type': 'application/json'}

    #cargas = '6:9,8:9'
    cargas = cargas.replace('/','"')
    carga =  '{' + cargas + '}'
    print(carga)
    d = json.loads(carga)
    pines = []
    def pedir_pin(institucion, grado, cantidad):

        for i in range(0, cantidad):
            payload_pj = {
                'institucion': institucion,
                'grado': grado,
                'creado_por': 'julian.ramirez',
                'a_o': '2020',
                'maxima_activacion': '2021-12-31',
                'maxima_usuario': '2021-12-31'
            }

            Datas = json.dumps(payload_pj)
            pedir = r.post('http://127.0.0.1:5000/api/pin', headers=HEADERS, data=Datas)
            print(pedir.text)
            guardar = str(pedir.text)
            pines.append(guardar)

    for key, value in d.items():
        pedir = pedir_pin(institucion, key, value)

    print(pines)
    def dividir_lista(alist, partes=1):
        length = len(alist)
        return [ alist[i*length // partes: (i+1)*length // partes]
                for i in range(partes) ]

    #el numero de partes debe ser equivalente a el "groupex" que va a ser el iterador en las listas
    M = dividir_lista(pines, partes=numarchivos)
    print(M[0])
    long = len(pines)
    groupex = numarchivos
    #M[i] es un grupo de 30 pines
    for i in range(0, groupex):
        print(i)
        shutil.copy2('maqueta_pin_TEST.xlsx','maq_pin({}).xlsx'.format(i))
        shutil.copy2('maqueta_consecutivo_TEST.xlsx','maq_consec({}).xlsx'.format(i))
        wbpin = openpyxl.load_workbook('maq_pin({})'.format(i)+ '.xlsx')

        wbconsec = openpyxl.load_workbook('maq_consec({})'.format(i)+ '.xlsx')


        sheet1 = wbpin.get_sheet_by_name("Base")

        sheet2 = wbconsec.get_sheet_by_name("Base")

        for iteracion, pin in enumerate(M[i]):
            c = 1 + 1
            cur1 = mysql.cursor()
            cur2 = mysql.cursor()
            cur3 = mysql.cursor()
            pin = pin.replace('"','')
            pin = str.rstrip(pin)
            cur1.execute('SELECT consecutivo FROM pines where pin = %s', (pin,))
            consecutivo = cur1.fetchall()
            consecutivo = ''.join(str(consecutivo))
            consecutivo = consecutivo.replace(',','').replace('(','').replace(')','')
            print(consecutivo)
            print(pin)
            cur2.execute('SELECT curso FROM pines where pin = %s', (pin,))
            cursopin = cur2.fetchone()
            cur3.execute('SELECT grado from cursos WHERE id =%s', (cursopin,))
            curso = cur3.fetchall()
            curso = ''.join(str(curso))
            curso = curso.replace(',','').replace('(','').replace(')','')
            curso = str(curso)
            curso = curso.replace("'","")
            print('course', curso)
            if curso == 'PJ':
                course = 'Prejardín'
            elif curso == 'J':
                course = 'Jardín'
            elif curso == 'T':
                course = 'Transición'
            elif curso == '1':
                course = 'Primero'
            elif curso == '2':
                course = 'Segundo'
            elif curso == '3':
                course = 'Tercero'
            elif curso == '4':
                course = 'Cuarto'
            elif curso == '5':
                course = 'Quinto'
            elif curso == '6':
                course = 'Sexto'
            elif curso == '7':
                course = 'Séptimo'
            elif curso == '8':
                course = 'Octavo'
            elif curso == '9':
                course = 'Noveno'
            elif curso == '10':
                course = 'Décimo'
            elif curso == '11':
                course = 'Undécimo'
            else:
                course = 'nada'
            print(course)
            sheet1.cell(row=iteracion + 1, column=1, value=course)
            sheet1.cell(row=iteracion + 1, column=2, value=pin)
            sheet1.cell(row=iteracion + 1, column=3, value='AVE')



            sheet2.cell(row=iteracion + 1, column=1, value=course)
            sheet2.cell(row=iteracion + 1, column=2, value=consecutivo)
            sheet2.cell(row=iteracion + 1, column=3, value='AVE')


            print(cur1._last_executed)
        nom_institucion = nom_institucion.replace('-',' ')
        sheet1.cell(row=32, column=1, value='Colegio:')
        sheet1.cell(row=32, column=2, value=nom_institucion)
        sheet1.cell(row=1, column=4, value=ciudad)
        print(pin)
        sheet2.cell(row=32, column=1, value='Colegio:')
        sheet2.cell(row=32, column=2, value=nom_institucion)
        sheet2.cell(row=1, column=4, value=ciudad)
        wbpin.save('maq_pin({})'.format(i)+ '.xlsx')
        wbconsec.save('maq_consec({})'.format(i)+ '.xlsx')
        response = 'Maquetas' + str(i)
        ts = calendar.timegm(time.gmtime())
        area=ctx.message.channel
        await ctx.send(response)
        await ctx.send(file=discord.File(r'C:\Users\HP\Desktop\bot_educar\maq_pin({}).xlsx'.format(i)))
        await ctx.send(file=discord.File(r'C:\Users\HP\Desktop\bot_educar\maq_consec({}).xlsx'.format(i)))
        shutil.move('maq_pin({}).xlsx'.format(i), 'maq_pin({})-{}-{}.xlsx'.format(i, institucion, ts))
        shutil.move('maq_consec({}).xlsx'.format(i), 'maq_consec({})-{}-{}.xlsx'.format(i, institucion, ts))

bot.run(TOKEN)
